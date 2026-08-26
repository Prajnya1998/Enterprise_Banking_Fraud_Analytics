"""
Enterprise Banking Fraud Analytics
-----------------------------------
Module: excel_formatter.py

Purpose:
Apply professional formatting to Excel reports.
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell


def format_workbook(workbook):
    """
    Apply professional formatting to all worksheets.

    Parameters:
        workbook: OpenPyXL workbook object.
    """

    # ==========================================================
    # HEADER FORMATTING
    # ==========================================================

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    center_alignment = Alignment(
        horizontal="center"
    )

    # ==========================================================
    # FORMAT EACH WORKSHEET
    # ==========================================================

    for sheet in workbook.worksheets:

        # ------------------------------------------------------
        # Format first-row headers
        # ------------------------------------------------------

        for cell in sheet[1]:

            if isinstance(cell, MergedCell):
                continue

            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment

        # ------------------------------------------------------
        # Apply borders to normal cells
        # ------------------------------------------------------

        for row in sheet.iter_rows():

            for cell in row:

                if isinstance(cell, MergedCell):
                    continue

                cell.border = thin_border

        # ------------------------------------------------------
        # Auto-adjust column widths
        # ------------------------------------------------------

        for column_cells in sheet.columns:

            normal_cells = [
                cell
                for cell in column_cells
                if not isinstance(cell, MergedCell)
            ]

            if not normal_cells:
                continue

            length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in normal_cells
            )

            column_letter = normal_cells[0].column_letter

            sheet.column_dimensions[
                column_letter
            ].width = length + 5