# ==========================================================
# Enterprise Banking Fraud Analytics
# Sprint 7 - Master Report
# ==========================================================

import pandas as pd

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

from openpyxl.utils import get_column_letter


# LOAD DATASET

dataset_path = "02_Data/Clean_Data/credit_card_transactions_clean.csv"

df = pd.read_csv(dataset_path)

print("=" * 45)
print("ENTERPRISE BANKING FRAUD ANALYTICS REPORT")
print("=" * 45)


# EXECUTIVE KPIs

total_transactions = len(df)

fraud_transactions = df["is_fraud"].sum()

genuine_transactions = total_transactions - fraud_transactions

fraud_rate = (fraud_transactions / total_transactions) * 100

average_amount = df["amt"].mean()

print(f"Total Transactions : {total_transactions:,}")
print(f"Genuine Transactions : {genuine_transactions:,}")
print(f"Fraud Transactions : {fraud_transactions:,}")
print(f"Fraud Rate : {fraud_rate:.2f}%")
print(f"Average Amount : ${average_amount:.2f}")


# TOP 5 FRAUD CATEGORIES

print("\n" + "=" * 25)
print("TOP 5 FRAUD CATEGORIES")
print("=" * 25)

fraud_category = (
    df[df["is_fraud"] == 1]
    .groupby("category")
    .size()
    .sort_values(ascending=False)
    .head(5)
)

print(fraud_category)

# TOP 5 FRAUD STATES

print("\n" + "=" * 20)
print("TOP 5 FRAUD STATES")
print("=" * 20)

fraud_state = (
    df[df["is_fraud"] == 1]
    .groupby("state")
    .size()
    .sort_values(ascending=False)
    .head(5)
)

print(fraud_state)


# TOP 5 FRAUD HOURS
df["trans_date_trans_time"] = pd.to_datetime(
    df["trans_date_trans_time"]
)


df["Hour"] = df["trans_date_trans_time"].dt.hour

print("\n" + "=" * 20)
print("TOP 5 FRAUD HOURS")
print("=" * 20)

fraud_hour = (
    df[df["is_fraud"] == 1]
    .groupby("Hour")
    .size()
    .sort_values(ascending=False)
    .head(5)
)

print(fraud_hour)


# KPI SUMMARY TABLE

kpi_summary = pd.DataFrame({

    "KPI": [
        "Total Transactions",
        "Genuine Transactions",
        "Fraud Transactions",
        "Fraud Rate (%)",
        "Average Transaction Amount ($)"
    ],

    "Value": [
        total_transactions,
        genuine_transactions,
        fraud_transactions,
        round(fraud_rate, 2),
        round(average_amount, 2)
    ]
})

print("\n" + "=" * 15)
print("KPI SUMMARY")
print("=" * 15)

print(kpi_summary)


# EXPORT COMPLETE REPORT

output_file = "09_Reports/Enterprise_Fraud_Report.xlsx"

with pd.ExcelWriter(
    output_file,
    engine="openpyxl"
) as writer:
 
    
    # Executive KPIs

    kpi_summary.to_excel(
        writer,
        sheet_name="Executive_KPIs",
        index=False
    )
    

    # Top Fraud Categories

    fraud_category.reset_index(
        name="Fraud_Count"
    ).to_excel(
        writer,
        sheet_name="Top_Fraud_Categories",
        index=False
    )


    # Top Fraud States

    fraud_state.reset_index(
        name="Fraud_Count"
    ).to_excel(
        writer,
        sheet_name="Top_Fraud_States",
        index=False
    )


    # Top Fraud Hours

    fraud_hour.reset_index(
        name="Fraud_Count"
    ).to_excel(
        writer,
        sheet_name="Top_Fraud_Hours",
        index=False
    )

    # FORMAT EXECUTIVE KPI SHEET

    worksheet = writer.sheets["Executive_KPIs"]

    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center"
    )


    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # BORDER STYLE

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # APPLY BORDERS

    for row in worksheet.iter_rows():

        for cell in row:
            cell.border = thin_border


    # AUTO ADJUST COLUMN WIDTH

    for column_cells in worksheet.columns:

        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )


        column_letter = column_cells[0].column_letter

        worksheet.column_dimensions[
            column_letter
        ].width = max_length + 3
    

    # FORMAT KPI VALUES

    for row in range(2, worksheet.max_row + 1):

        kpi_name = worksheet[f"A{row}"].value
        value_cell = worksheet[f"B{row}"]

        if kpi_name == "Fraud Rate (%)":
            value_cell.number_format = "0.00%"

        elif kpi_name == "Average Transaction Amount ($)":
            value_cell.number_format = "$#,##0.00"

        else:
            value_cell.number_format = "#,##0"


    # HIGHLIGHT RISK KPIs

    risk_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    risk_font = Font(
        bold=True,
        color="9C0006"
    )

    for row in range(2, worksheet.max_row + 1):

        kpi_name = worksheet[f"A{row}"].value

        if kpi_name in [
            "Fraud Transactions",
            "Fraud Rate (%)"
        ]:
        
            worksheet[f"A{row}"].fill = risk_fill
            worksheet[f"B{row}"].fill = risk_fill

            worksheet[f"A{row}"].font = risk_font
            worksheet[f"B{row}"].font = risk_font

    
    # CREATE DASHBOARD SHEET

    dashboard = writer.book.create_sheet("Dashboard")

    dashboard["A1"] = "ENTERPRISE FRAUD DASHBOARD"

    dashboard["A1"].font = Font(
        size=18,
        bold=True,
        color="FFFFFF"
    )

    dashboard["A1"].fill =PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    dashboard.merge_cells("A1:F1")
    dashboard["A1"].alignment = Alignment(horizontal="center")

    # KPI Labels

    dashboard["A3"] = "Total Transactions"
    dashboard["A4"] = "Fraud Transactions"
    dashboard["A5"] = "Fraud Rate"
    dashboard["A6"] = "Average Amount"

    # KPI Values

    dashboard["B3"] = total_transactions
    dashboard["B4"] = fraud_transactions
    dashboard["B5"] = f"{fraud_rate}%"
    dashboard["B6"] = average_amount

    dashboard["B3"].number_format = "#,##0"
    dashboard["B4"].number_format = "#,##0"
    dashboard["B6"].number_format = "$#,##0.00"

    
    dashboard["A9"] = "Executive Summary"

    dashboard["A9"].font = Font(
        bold=True,
        size=14
    )

    dashboard["A11"] = (
        "• Fraud Rate remains below 1%."
    )

    dashboard["A12"] = (
        "• Grocery POS contributes the highest fraud volume."
    )

    dashboard["A13"] = (
        "• Most fraud occurs between 10 PM and 11 PM."
    )

    dashboard["A14"] = (
        "• New York records the highest fraud count."
    )

    # FREEZE DASHBOARD

    dashboard.freeze_panes = "A3"

    dashboard["A3"] = "Total Transactions"
    dashboard["A4"] = "Fraud Transactions"
    dashboard["A5"] = "Fraud Rate"
    dashboard["A6"] = "Average Amount"

    # FORMAT KPI LABELS

    for cell in ["A3", "A4", "A5", "A6"]:
        dashboard[cell].font = Font(
            bold=True,
            size=12
        )

    # FORMAT KPI VALUES

    for cell in ["B3", "B4", "B5", "B6"]:
        dashboard[cell].font = Font(
            bold=True,
            size=14
        )

    # AUTO ADJUST DASHBOARD COLUMNS

    for column in ["A", "B"]:
        max_length = 0
        for cell in dashboard[column]:
            if cell.value is not None:
                max_length = max(
                    max_length, 
                    len(str(cell.value))
                )
        dashboard.column_dimensions[column_letter].width = max_length + 5


print("\n" + "=" * 40)
print("ENTERPRISE REPORT CREATED SUCCESSFULLY")
print("=" * 40)

print(f"Location : {output_file}")


