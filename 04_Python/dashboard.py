"""
Enterprise Banking Fraud Analytics
-----------------------------------
Module : dashboard.py

Purpose:
Create a professional executive fraud dashboard.
"""

from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image


def create_dashboard(workbook, kpis):
    """
    Create the executive fraud dashboard.

    Parameters:
        workbook: OpenPyXL workbook object.
        kpis (dict): Executive KPI values.

    Returns:
        Dashboard worksheet.
    """

    # ==========================================================
    # CREATE DASHBOARD
    # ==========================================================

    dashboard = workbook.create_sheet("Dashboard")

    # ==========================================================
    # DASHBOARD TITLE
    # ==========================================================

    dashboard["A1"] = "ENTERPRISE BANKING FRAUD ANALYTICS"

    dashboard["A1"].font = Font(
        bold=True,
        size=20,
        color="FFFFFF"
    )

    dashboard["A1"].fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    dashboard["A1"].alignment = Alignment(
        horizontal="center"
    )

    # ==========================================================
    # COLUMN WIDTHS
    # ==========================================================

    dashboard.column_dimensions["A"].width = 24
    dashboard.column_dimensions["B"].width = 18
    dashboard.column_dimensions["C"].width = 4
    dashboard.column_dimensions["D"].width = 24
    dashboard.column_dimensions["E"].width = 18
    dashboard.column_dimensions["F"].width = 4
    dashboard.column_dimensions["G"].width = 24
    dashboard.column_dimensions["H"].width = 18
    dashboard.column_dimensions["I"].width = 4
    dashboard.column_dimensions["J"].width = 24
    dashboard.column_dimensions["K"].width = 18

    # ==========================================================
    # KPI SECTION TITLE
    # ==========================================================

    dashboard["A3"] = "EXECUTIVE KPIs"

    dashboard["A3"].font = Font(
        bold=True,
        size=14
    )

    # ==========================================================
    # KPI CARDS
    # ==========================================================

    kpi_cards = [

        ("A5", "B5", "Total Transactions", kpis["Total Transactions"]),

        ("D5", "E5", "Fraud Transactions", kpis["Fraud Transactions"]),

        ("G5", "H5", "Fraud Rate", kpis["Fraud Rate"] / 100),

        ("J5", "K5", "Total Fraud Amount", kpis["Total Fraud Amount"])

    ]

    card_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAF7",
        end_color="D9EAF7"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for label_cell, value_cell, label, value in kpi_cards:

        dashboard[label_cell] = label

        dashboard[label_cell].font = Font(
            bold=True,
            size=11
        )

        dashboard[label_cell].fill = card_fill

        dashboard[label_cell].alignment = Alignment(
            horizontal="center"
        )

        dashboard[label_cell].border = border

        dashboard[value_cell] = value

        dashboard[value_cell].font = Font(
            bold=True,
            size=15
        )

        dashboard[value_cell].fill = card_fill

        dashboard[value_cell].alignment = Alignment(
            horizontal="center"
        )

        dashboard[value_cell].border = border

    # ==========================================================
    # KPI NUMBER FORMATTING
    # ==========================================================

    dashboard["B5"].number_format = "#,##0"
    dashboard["E5"].number_format = "#,##0"
    dashboard["H5"].number_format = "0.00%"
    dashboard["K5"].number_format = "$#,##0.00"

    # ==========================================================
    # SECONDARY KPI SECTION
    # ==========================================================

    dashboard["A8"] = "Additional Risk Indicators"

    dashboard["A8"].font = Font(
        bold=True,
        size=13
    )

    dashboard["A10"] = "Genuine Transactions"
    dashboard["B10"] = kpis["Genuine Transactions"]

    dashboard["D10"] = "Average Amount"
    dashboard["E10"] = kpis["Average Amount"]

    dashboard["G10"] = "Maximum Amount"
    dashboard["H10"] = kpis["Maximum Amount"]

    dashboard["A10"].font = Font(bold=True)
    dashboard["D10"].font = Font(bold=True)
    dashboard["G10"].font = Font(bold=True)

    dashboard["B10"].number_format = "#,##0"
    dashboard["E10"].number_format = "$#,##0.00"
    dashboard["H10"].number_format = "$#,##0.00"

    # ==========================================================
    # CHART SECTION
    # ==========================================================

    dashboard["A13"] = "FRAUD ANALYSIS"

    dashboard["A13"].font = Font(
        bold=True,
        size=14
    )

    # ==========================================================
    # CHART FILE LOCATIONS
    # ==========================================================

    image_directory = Path("images")

    category_chart = image_directory / "Fraud_By_Category.png"
    state_chart = image_directory / "Fraud_By_State.png"
    hour_chart = image_directory / "Fraud_By_Hour.png"

    # ==========================================================
    # FRAUD BY CATEGORY
    # ==========================================================

    if category_chart.exists():

        category_image = Image(
            str(category_chart)
        )

        category_image.width = 500
        category_image.height = 300

        dashboard.add_image(
            category_image,
            "A15"
        )

    # ==========================================================
    # FRAUD BY STATE
    # ==========================================================

    if state_chart.exists():

        state_image = Image(
            str(state_chart)
        )

        state_image.width = 500
        state_image.height = 300

        dashboard.add_image(
            state_image,
            "J15"
        )

    # ==========================================================
    # FRAUD BY HOUR
    # ==========================================================

    if hour_chart.exists():

        hour_image = Image(
            str(hour_chart)
        )

        hour_image.width = 850
        hour_image.height = 400

        dashboard.add_image(
            hour_image,
            "A32"
        )

    # ==========================================================
    # FREEZE PANES
    # ==========================================================

    dashboard.freeze_panes = "A4"

    # ==========================================================
    # ROW HEIGHTS
    # ==========================================================

    dashboard.row_dimensions[1].height = 30
    dashboard.row_dimensions[5].height = 25
    dashboard.row_dimensions[10].height = 22

    return dashboard